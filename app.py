from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import os, datetime, json, openpyxl
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "amman_studio_secret_2024"

# Use PostgreSQL on Render, SQLite locally
database_url = os.environ.get("DATABASE_URL", "sqlite:///amman_studio.db")
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True, "pool_recycle": 300}
app.config["UPLOAD_FOLDER"] = os.path.join("static", "images")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024
app.config["RAZORPAY_KEY_ID"] = "rzp_test_YourKeyHere"
app.config["RAZORPAY_KEY_SECRET"] = "YourSecretHere"
EXCEL_FILE = os.path.join("instance", "amman_data.xlsx")

db = SQLAlchemy(app)
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}

# ─── Models ───────────────────────────────────────────────
class Service(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text, nullable=False)
    price = db.Column(db.Float, nullable=False)
    category = db.Column(db.String(50), nullable=False)
    image = db.Column(db.String(200), default="default.jpg")
    featured = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class Booking(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    service_id = db.Column(db.Integer, db.ForeignKey("service.id"))
    service = db.relationship("Service", backref="bookings")
    message = db.Column(db.Text)
    status = db.Column(db.String(20), default="pending")
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class Contact(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), nullable=False)
    subject = db.Column(db.String(200))
    message = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class Review(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    rating = db.Column(db.Integer, nullable=False)
    comment = db.Column(db.Text, nullable=False)
    service_id = db.Column(db.Integer, db.ForeignKey("service.id"))
    service = db.relationship("Service", backref="reviews")
    approved = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)

# ─── Helpers ──────────────────────────────────────────────
# ─── Excel Helper ─────────────────────────────────────────
def save_to_excel(sheet_name, headers, row_data):
    os.makedirs("instance", exist_ok=True)
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
    else:
        ws = wb[sheet_name]
    ws.append(row_data)
    wb.save(EXCEL_FILE)

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated

# ─── Frontend Routes ──────────────────────────────────────
@app.route("/")
def index():
    featured = Service.query.filter_by(featured=True).limit(6).all()
    reviews = Review.query.filter_by(approved=True).order_by(Review.created_at.desc()).limit(6).all()
    return render_template("index.html", featured=featured, reviews=reviews)

@app.route("/services")
def services():
    category = request.args.get("category", "all")
    if category == "all":
        all_services = Service.query.all()
    else:
        all_services = Service.query.filter_by(category=category).all()
    categories = db.session.query(Service.category).distinct().all()
    categories = [c[0] for c in categories]
    return render_template("services.html", services=all_services, categories=categories, active=category)

@app.route("/service/<int:id>")
def service_detail(id):
    svc = Service.query.get_or_404(id)
    reviews = Review.query.filter_by(service_id=id, approved=True).all()
    return render_template("service_detail.html", service=svc, reviews=reviews)

@app.route("/help")
def help_center():
    return render_template("help.html")

@app.route("/booking", methods=["GET", "POST"])
def booking():
    services = Service.query.all()
    if request.method == "POST":
        b = Booking(
            name=request.form["name"],
            email=request.form["email"],
            phone=request.form["phone"],
            service_id=request.form["service_id"],
            message=request.form.get("message", "")
        )
        db.session.add(b)
        db.session.commit()
        service_name = b.service.name if b.service else ""
        save_to_excel("Bookings",
            ["ID", "Name", "Email", "Phone", "Service", "Message", "Status", "Date"],
            [b.id, b.name, b.email, b.phone, service_name, b.message, b.status, str(b.created_at)]
        )
        return jsonify({"success": True, "message": "Booking confirmed! We'll contact you soon."})
    return render_template("booking.html", services=services)

@app.route("/contact", methods=["GET", "POST"])
def contact():
    if request.method == "POST":
        c = Contact(
            name=request.form["name"],
            email=request.form["email"],
            subject=request.form.get("subject", ""),
            message=request.form["message"]
        )
        db.session.add(c)
        db.session.commit()
        save_to_excel("Contacts",
            ["ID", "Name", "Email", "Subject", "Message", "Date"],
            [c.id, c.name, c.email, c.subject, c.message, str(c.created_at)]
        )
        return jsonify({"success": True, "message": "Message sent! We'll reply within 24 hours."})
    return render_template("contact.html")

@app.route("/review", methods=["POST"])
def add_review():
    r = Review(
        name=request.form["name"],
        rating=int(request.form["rating"]),
        comment=request.form["comment"],
        service_id=request.form.get("service_id")
    )
    db.session.add(r)
    db.session.commit()
    service_name = r.service.name if r.service else ""
    save_to_excel("Reviews",
        ["ID", "Name", "Rating", "Comment", "Service", "Date"],
        [r.id, r.name, r.rating, r.comment, service_name, str(r.created_at)]
    )
    return jsonify({"success": True, "message": "Thank you for your review!"})

# ─── Payment Routes ──────────────────────────────────────
@app.route("/payment/initiate", methods=["POST"])
def initiate_payment():
    service_id = request.form.get("service_id")
    s = Service.query.get(service_id)
    if not s:
        return jsonify({"success": False, "message": "Service not found"})
    return jsonify({
        "success": True,
        "key": app.config["RAZORPAY_KEY_ID"],
        "amount": int(s.price * 100),
        "service_name": s.name,
        "currency": "INR"
    })

@app.route("/payment/verify", methods=["POST"])
def verify_payment():
    data = request.get_json()
    b = Booking(
        name=data.get("name"),
        email=data.get("email"),
        phone=data.get("phone"),
        service_id=data.get("service_id"),
        message=data.get("message", ""),
        status="paid"
    )
    db.session.add(b)
    db.session.commit()
    service_name = b.service.name if b.service else ""
    save_to_excel("Bookings",
        ["ID", "Name", "Email", "Phone", "Service", "Message", "Status", "Payment ID", "Date"],
        [b.id, b.name, b.email, b.phone, service_name, b.message, "paid", data.get("razorpay_payment_id", ""), str(b.created_at)]
    )
    return jsonify({"success": True})

# ─── API ──────────────────────────────────────────────────
@app.route("/api/services")
def api_services():
    services = Service.query.all()
    return jsonify([{
        "id": s.id, "name": s.name, "description": s.description,
        "price": s.price, "category": s.category, "image": s.image, "featured": s.featured
    } for s in services])

# ─── Export Routes ───────────────────────────────────────
@app.route("/admin/export/excel")
@admin_required
def export_excel():
    from flask import send_file
    import io
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Bookings sheet
    ws1 = wb.create_sheet("Bookings")
    ws1.append(["ID", "Name", "Email", "Phone", "Service", "Message", "Status", "Date"])
    for b in Booking.query.order_by(Booking.created_at.desc()).all():
        ws1.append([b.id, b.name, b.email, b.phone,
                    b.service.name if b.service else "",
                    b.message or "", b.status, str(b.created_at)])

    # Contacts sheet
    ws2 = wb.create_sheet("Contacts")
    ws2.append(["ID", "Name", "Email", "Subject", "Message", "Date"])
    for c in Contact.query.order_by(Contact.created_at.desc()).all():
        ws2.append([c.id, c.name, c.email, c.subject or "", c.message, str(c.created_at)])

    # Reviews sheet
    ws3 = wb.create_sheet("Reviews")
    ws3.append(["ID", "Name", "Rating", "Comment", "Service", "Date"])
    for r in Review.query.order_by(Review.created_at.desc()).all():
        ws3.append([r.id, r.name, r.rating, r.comment,
                    r.service.name if r.service else "", str(r.created_at)])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="amman_studio_data.xlsx",
                     as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── Admin Routes ─────────────────────────────────────────
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        admin = Admin.query.filter_by(username=request.form["username"]).first()
        if admin and check_password_hash(admin.password, request.form["password"]):
            session["admin"] = True
            return redirect(url_for("admin_dashboard"))
        return render_template("admin/login.html", error="Invalid credentials")
    return render_template("admin/login.html")

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    return redirect(url_for("admin_login"))

@app.route("/admin")
@admin_required
def admin_dashboard():
    stats = {
        "services": Service.query.count(),
        "bookings": Booking.query.count(),
        "contacts": Contact.query.count(),
        "reviews": Review.query.count(),
        "pending": Booking.query.filter_by(status="pending").count()
    }
    recent_bookings = Booking.query.order_by(Booking.created_at.desc()).limit(5).all()
    return render_template("admin/dashboard.html", stats=stats, recent_bookings=recent_bookings)

@app.route("/admin/services", methods=["GET", "POST"])
@admin_required
def admin_services():
    if request.method == "POST":
        image_filename = "default.jpg"
        if "image" in request.files:
            file = request.files["image"]
            if file and allowed_file(file.filename):
                image_filename = secure_filename(file.filename)
                file.save(os.path.join(app.config["UPLOAD_FOLDER"], image_filename))
        s = Service(
            name=request.form["name"],
            description=request.form["description"],
            price=float(request.form["price"]),
            category=request.form["category"],
            image=image_filename,
            featured="featured" in request.form
        )
        db.session.add(s)
        db.session.commit()
        return redirect(url_for("admin_services"))
    services = Service.query.all()
    return render_template("admin/services.html", services=services)

@app.route("/admin/services/delete/<int:id>")
@admin_required
def delete_service(id):
    try:
        s = Service.query.get_or_404(id)
        db.session.delete(s)
        db.session.commit()
    except Exception:
        db.session.rollback()
    return redirect(url_for("admin_services"))

@app.route("/admin/services/edit/<int:id>", methods=["GET", "POST"])
@admin_required
def edit_service(id):
    s = Service.query.get_or_404(id)
    if request.method == "POST":
        try:
            s.name = request.form.get("name", s.name)
            s.price = float(request.form.get("price", s.price))
            s.description = request.form.get("description", s.description)
            s.category = request.form.get("category", s.category)
            s.featured = request.form.get("featured") == "on"
            db.session.commit()
        except Exception:
            db.session.rollback()
        return redirect(url_for("admin_services"))
    return redirect(url_for("admin_services"))

@app.route("/admin/services/update-price/<int:id>", methods=["POST"])
@admin_required
def update_price(id):
    s = Service.query.get_or_404(id)
    s.price = float(request.form.get("price", s.price))
    db.session.commit()
    return jsonify({"success": True, "price": s.price})

@app.route("/admin/bookings")
@admin_required
def admin_bookings():
    bookings = Booking.query.order_by(Booking.created_at.desc()).all()
    return render_template("admin/bookings.html", bookings=bookings)

@app.route("/admin/bookings/status/<int:id>/<status>")
@admin_required
def update_booking_status(id, status):
    try:
        b = Booking.query.get_or_404(id)
        b.status = status
        db.session.commit()
    except Exception as e:
        db.session.rollback()
    return redirect(url_for("admin_bookings"))

@app.route("/admin/contacts")
@admin_required
def admin_contacts():
    contacts = Contact.query.order_by(Contact.created_at.desc()).all()
    return render_template("admin/contacts.html", contacts=contacts)

@app.route("/admin/reviews")
@admin_required
def admin_reviews():
    reviews = Review.query.order_by(Review.created_at.desc()).all()
    return render_template("admin/reviews.html", reviews=reviews)

# ─── Init ─────────────────────────────────────────────────
def init_db():
    with app.app_context():
        db.create_all()
        if not Admin.query.first():
            db.session.add(Admin(username="admin", password=generate_password_hash("amman2024")))
        if not Review.query.first():
            sample_reviews = [
                Review(name="Rahul Sharma", rating=5, comment="Amazing work! Amman Studio delivered exactly what I wanted. Highly recommended!", approved=True),
                Review(name="Priya Nair", rating=5, comment="The best design service. Professional team, beautiful results!", approved=True),
                Review(name="Arjun Patel", rating=5, comment="Logo design was perfect. They understood my brand vision completely. Great team!", approved=True),
                Review(name="Meena Krishnan", rating=5, comment="PPT presentation was stunning! Got great feedback from my professors.", approved=True),
                Review(name="Vikram Das", rating=5, comment="Website design exceeded my expectations. Very professional and fast delivery!", approved=True),
                Review(name="Sneha Reddy", rating=5, comment="Word document was neatly formatted and delivered on time. Will use again!", approved=True),
            ]
            db.session.add_all(sample_reviews)
        if not Service.query.first():
            sample_services = [
                # Design
                Service(name="Logo Design", description="Creative and unique logo design for your brand — minimal, modern, and memorable.", price=999, category="Design", image="https://images.unsplash.com/photo-1626785774573-4b799315345d?w=400&q=80", featured=True),
                # Web
                Service(name="Website Design", description="Modern, fully responsive website design and development for your business.", price=4999, category="Web", image="https://images.unsplash.com/photo-1547658719-da2b51169166?w=400&q=80", featured=True),
                Service(name="Webpage Design", description="Single page professional web design with clean layout and smooth animations.", price=999, category="Web", image="https://images.unsplash.com/photo-1581291518857-4e27b48ff24e?w=400&q=80", featured=False),
                # Presentation
                Service(name="PPT Presentation", description="Creative and visually stunning PowerPoint presentations for any occasion.", price=499, category="Presentation", image="https://images.unsplash.com/photo-1508921912186-1d1a45ebb3c1?w=400&q=80", featured=True),
                Service(name="Word Presentation", description="Professionally formatted Word documents and reports with creative layouts.", price=399, category="Presentation", image="https://images.unsplash.com/photo-1455390582262-044cdead277a?w=400&q=80", featured=False),
                Service(name="Project Presentation", description="Stunning project presentations with creative visuals, charts, and structured storytelling.", price=1999, category="Presentation", image="https://images.unsplash.com/photo-1552664730-d307ca884978?w=400&q=80", featured=True),
            ]
            db.session.add_all(sample_services)
        db.session.commit()

os.makedirs(os.path.join("static", "images"), exist_ok=True)
init_db()

if __name__ == "__main__":
    print("\nAmman Studio is running!")
    print("  Website  -> http://localhost:5000")
    print("  Admin    -> http://localhost:5000/admin/login")
    print("  Login    -> admin / amman2024\n")
    app.run(debug=True, port=5000)
